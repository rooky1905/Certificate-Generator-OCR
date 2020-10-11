import cv2
import openpyxl 
# from datetime import date
from zipfile import ZipFile
import os
import pytesseract #optical char recogn
# import matplotlib.pyplot as plt
from pytesseract import Output

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe'

zipObj = ZipFile('sample.zip', 'w')
template_path = 'template12.jpeg'
details_path = 'names.xlsx'
output_path = 'D:\Python Everything\Certificate_Sukku'
 
font_size = 3
font_color = (0,0,0) 
   
obj = openpyxl.load_workbook(details_path) 
sheet = obj.active
size = sheet.max_row
print(size)
img = cv2.imread(template_path)

data = pytesseract.image_to_data(img, output_type=Output.DICT)

for i in range (len(data['text'])):
    data['text'][i] = data['text'][i].lower()

ind_to = data['text'].index('to')

ind_this = data['text'].index('this')
if ind_this == -1:
    ind_this = data['text'].index('is')

ind_for = data['text'].index('in')
if ind_for == -1:
    ind_for = data['text'].index('for')
    if ind_for == -1:
        ind_for = data['text'].index('sample')

ind_date = data['text'].index('date')

left = (data['left'][ind_to] + data['left'][ind_this])//2
right = (data['top'][ind_to] + data['left'][ind_for])//2

left_date = data['left'][ind_date]
right_date = data['top'][ind_date]

for i in range(1,size+1): 
    # grabs the row=i and column=1 cell  
    # that contains the name value of that 
    # cell is stored in the variable certi_name 
    get_name = sheet.cell(row = i ,column = 1)
    get_date = sheet.cell(row=i, column = 2)
    certi_name = get_name.value
    get_date = get_date.value
    # print(type(get_date))
    #print(certi_name)
    # read the certificate template 
    img = cv2.imread(template_path)
    # choose the font from opencv 
    font = cv2.FONT_HERSHEY_PLAIN               

    cv2.putText(img, certi_name, 
              (left-50,right+40),  
              font, 
              font_size, 
              font_color, 4)

    cv2.putText(img, get_date, 
              (left_date - 20, right_date - 20),  
              font, 
              font_size-2, 
              font_color, 2)

    # Output path along with the name of the 
    # certificate generated 
    certi_path = output_path + '/certi{}'.format(i) + '.png'
    img_path = 'certi{}'.format(i) + '.png'
    # Save the certificate
               
    cv2.imwrite(certi_path,img)
    zipObj.write(img_path)
    os.remove(img_path)
    
zipObj.close()


