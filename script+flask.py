from flask import Flask, render_template, request, redirect, url_for, session, send_file
from werkzeug.utils import secure_filename
import cv2
import openpyxl 
from zipfile import ZipFile
import os
import pytesseract 
from pytesseract import Output
app = Flask(__name__)

@app.route('/')
def upload():
   return render_template('index.html')

@app.route('/uploader', methods = ['GET', 'POST'])
def upld_file():
    if request.method == 'POST':
        certificate = request.files['certi']
        participants = request.files['xl']
        certificate.save(secure_filename(certificate.filename))
        participants.save(secure_filename(participants.filename))
        pytesseract.pytesseract.tesseract_cmd = 'C:/Users/Sandali/AppData/Local/Programs/Python/Python38-32/Lib/site-packages/Tesseract-OCR/tesseract.exe'
        zipObj = ZipFile('sample.zip', 'w')
        template_path = certificate.filename
        details_path = participants.filename
        output_path = 'C:/Users/Sandali/Desktop/VitHack'
        font_size = 3
        font_color = (0,0,0) 

            # loading the details.xlsx workbook  
            # and grabbing the active sheet 
        obj = openpyxl.load_workbook(details_path) 
        sheet = obj.active
        size = sheet.max_row
        img = cv2.imread(template_path)

        data = pytesseract.image_to_data(img, output_type=Output.DICT)
        for i in range (len(data['text'])):
            data['text'][i] = data['text'][i].lower()
            
        
        ind_to = data['text'].index('to')
        
        try:
            ind_this = data['text'].index('this')
        except:
            ind_this = data['text'].index('is')
        
        try:
            ind_for = data['text'].index('in')
        except:
            try:
                ind_for = data['text'].index('for')
            except:
                try:
                    ind_for = data['text'].index('lorem')
                except:
                    ind_for = data['text'].index('sample')
       
        ind_date = data['text'].index('date')
            
        left = (data['left'][ind_to] + data['left'][ind_this])//2
        right = (data['top'][ind_to] + data['left'][ind_for])//2
        
        left_date = data['left'][ind_date]
        right_date = data['top'][ind_date]

            # printing for the first 10 names in the 
            # excel sheet 
        for i in range(1,size+1): 
                # grabs the row=i and column=1 cell  
                # that contains the name value of that 
                # cell is stored in the variable certi_name 
            get_name = sheet.cell(row = i ,column = 1) 
            certi_name = get_name.value 
            get_date = sheet.cell(row = i, column = 2)
            get_date = get_date.value
                # read the certificate template 
            img = cv2.imread(template_path)
                # choose the font from opencv 
            font = cv2.FONT_HERSHEY_PLAIN               
            cv2.putText(img, certi_name,(left-55,right+35),font,font_size,font_color, 3)
            cv2.putText(img, get_date,(left_date - 20, right_date - 20),font,1,font_color, 2)

                # Output path along with the name of the 
                # certificate generated 
            certi_path = output_path + '/certi{}'.format(i) + '.png'
            img_path = 'certi{}'.format(i) + '.png'
                # Save the certificate
                           
            cv2.imwrite(certi_path,img)
            zipObj.write(img_path)
            os.remove(img_path)
                
        zipObj.close()
        return render_template('result.html')

@app.route('/download')
def download_file():
    p = 'sample.zip'
    return send_file(p, as_attachment = True)

if (__name__=='__main__'):
    app.run(debug=True)